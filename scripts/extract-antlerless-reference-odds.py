from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook


RAW_XLSX = Path(r"C:\UOGA HUNTS\raw_data_2026\25_antlerless_drawing_odds_report.xlsx")
OUT_CSV = Path(r"C:\UOGA HUNTS\raw_data_2026\antlerless_reference_odds_2025.csv")

HUNT_RE = re.compile(r"Hunt:\s*([A-Z]{2}\d{4})\s+(.*)")


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def to_int(value) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def ratio_to_pct(text: str) -> float | None:
    text = normalize_text(text)
    if not text or text.upper() == "N/A":
        return None
    if " in " in text:
        left, right = text.split(" in ", 1)
        try:
            return float(left.strip()) / float(right.strip()) * 100.0
        except ValueError:
            return None
    return None


def main() -> None:
    wb = load_workbook(RAW_XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_out: list[dict[str, object]] = []
    current_code = ""
    current_name = ""

    for row in ws.iter_rows(values_only=True):
        values = [normalize_text(v) for v in row]
        joined = " ".join(v for v in values if v)
        hunt_match = HUNT_RE.search(joined)
        if hunt_match:
            current_code = hunt_match.group(1)
            current_name = hunt_match.group(2).strip()
            continue

        if not current_code:
            continue

        # resident columns: 0,1,3,4,5,6
        resident_point = to_int(row[0] if len(row) > 0 else None)
        if resident_point is not None:
            rows_out.append(
                {
                    "hunt_code": current_code,
                    "hunt_name": current_name,
                    "residency": "Resident",
                    "points": resident_point,
                    "applicants": to_int(row[1] if len(row) > 1 else None) or 0,
                    "bonus_permits": to_int(row[3] if len(row) > 3 else None) or 0,
                    "regular_permits": to_int(row[4] if len(row) > 4 else None) or 0,
                    "total_permits": to_int(row[5] if len(row) > 5 else None) or 0,
                    "success_ratio_text": normalize_text(row[6] if len(row) > 6 else None),
                    "success_pct": ratio_to_pct(normalize_text(row[6] if len(row) > 6 else None)),
                    "source_workbook": RAW_XLSX.name,
                }
            )

        # nonresident columns: 7,9,11,12,13,14
        nonresident_point = to_int(row[7] if len(row) > 7 else None)
        if nonresident_point is not None:
            rows_out.append(
                {
                    "hunt_code": current_code,
                    "hunt_name": current_name,
                    "residency": "Nonresident",
                    "points": nonresident_point,
                    "applicants": to_int(row[9] if len(row) > 9 else None) or 0,
                    "bonus_permits": to_int(row[11] if len(row) > 11 else None) or 0,
                    "regular_permits": to_int(row[12] if len(row) > 12 else None) or 0,
                    "total_permits": to_int(row[13] if len(row) > 13 else None) or 0,
                    "success_ratio_text": normalize_text(row[14] if len(row) > 14 else None),
                    "success_pct": ratio_to_pct(normalize_text(row[14] if len(row) > 14 else None)),
                    "source_workbook": RAW_XLSX.name,
                }
            )

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "hunt_code",
                "hunt_name",
                "residency",
                "points",
                "applicants",
                "bonus_permits",
                "regular_permits",
                "total_permits",
                "success_ratio_text",
                "success_pct",
                "source_workbook",
            ],
        )
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {OUT_CSV}")


if __name__ == "__main__":
    main()
