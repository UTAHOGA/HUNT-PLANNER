from __future__ import annotations

import csv
import json
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
try:
    from pypdf import PdfReader
except ModuleNotFoundError:
    from PyPDF2 import PdfReader


WORKSPACE_ROOT = Path(r"C:\DOWNLOADS\test website\HUNT-PLANNER")
UOGA_ROOT = Path(r"C:\UOGA HUNTS")
PDF_PATH = Path(r"C:\DOWNLOADS\2026-04-rac-packet (1).pdf")
RAC_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "2026-04-rac-packet.xlsx"
RAC_WORKBOOK_ALT_PATH = UOGA_ROOT / "raw_data_2026" / "2026-04-rac-packet a.xlsx"
DEER_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "BUCK DEER.xlsx"
TURKEY_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "Turkey Bearded.xlsx"
TURKEY_EITHER_SEX_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "Turkey  Either Sex.xlsx"
ANTLERLESS_ELK_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "antlerless elk.xlsx"
ANTLERLESS_DEER_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "antlerless deer.xlsx"
BULL_MOOSE_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "bull moose.xlsx"
COW_MOOSE_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "cow moose.xlsx"
MOUNTAIN_GOAT_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "mtn goat.xlsx"
PRONGHORN_BUCK_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "pronghorn buck.xlsx"
PRONGHORN_DOE_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "pronghorn doe.xlsx"
BISON_BULL_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "bison bull.xlsx"
BISON_COW_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "bison  cow.xlsx"
DESERT_SHEEP_RAM_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "desert sheep ram.xlsx"
ROCKY_RAM_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "rocky mtn. sheep.xlsx"
ROCKY_EWE_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "rocky mtn sheep ewe.xlsx"
DEER_HUNTERS_CHOICE_WORKBOOK_PATH = UOGA_ROOT / "raw_data_2026" / "deer hunter's choice.xlsx"
OUTPUT_DIR = WORKSPACE_ROOT / "data" / "uoga_recommended_permits_2026"
PROCESSED_OUTPUT_PATH = UOGA_ROOT / "processed_data" / "recommended_permits_2026.csv"
PROCESSED_YEAR_OUTPUT_PATH = UOGA_ROOT / "processed_data" / "2026" / "recommended_permits_2026.csv"

RAW_SECTION_PATH = OUTPUT_DIR / "layer_04_recommended_permits_raw_2026.jsonl"
CSV_PATH = OUTPUT_DIR / "recommended_permits_2026.csv"
VALIDATION_PATH = OUTPUT_DIR / "recommended_permits_2026.validation.json"
ACCEPTANCE_PATH = OUTPUT_DIR / "recommended_permits_2026.acceptance.json"
STATUS_PATH = OUTPUT_DIR / "recommended_permits_2026.status.json"
SOURCE_MANIFEST_PATH = OUTPUT_DIR / "layer_01_source_manifest_2026_recommended_permits.json"

HUNT_CODE_RE = re.compile(r"\b([A-Z]{2}\d{4})\b")
YEAR_MARKER_RE = re.compile(r"2025\s+Permits\s+2026\s+Permi\w*", re.IGNORECASE)
DASHES = {"–", "-", "—"}

WEAPON_OPTIONS = [
    "Archery , Muzzleloader , Shotgun",
    "Any Bull/Hunter's Choice (any weapon)",
    "Any Legal Weapon",
    "September Archery",
    "Late Archery",
    "Muzzleloader",
    "Multiseason",
    "Archery",
    "HAMSS",
]
SEX_TYPE_OPTIONS = [
    "Hunter's Choice",
    "Hunter ’s Choice",
    "Female Only",
]


@dataclass
class SectionContext:
    page_number: int
    page_heading: str
    section_title: str
    header_type: str


def clean_text(value: str) -> str:
    value = value.replace("\x00", " ")
    value = value.replace("W eapon", "Weapon")
    value = value.replace("WeaponHunt", "Weapon Hunt")
    value = value.replace("T ype", "Type")
    value = value.replace("Clif fs", "Cliffs")
    value = value.replace("V alley", "Valley")
    value = value.replace("Mtns/Central Mtns", "Wasatch Mtns/Central Mtns")
    value = value.replace("Mtn/V ernal", "Mtn/Vernal")
    value = value.replace("Bitter\nCreek", "Bitter Creek")
    value = value.replace("Creek/SouthAny", "Creek/South Any")
    value = value.replace("PromontoryAny", "Promontory Any")
    value = value.replace("SouthAny", "South Any")
    value = value.replace("BenchAny", "Bench Any")
    value = value.replace("LEAny", "LE Any")
    value = value.replace("T impanogos", "Timpanogos")
    value = value.replace("W est", "West")
    value = value.replace("T riangle", "Triangle")
    value = value.replace("P ariette", "Pariette")
    value = value.replace("ResNon", "Res Non")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"\b([A-Z]{2}\d{2})\s+(\d{2})\b", r"\1\2", value)
    value = re.sub(r"\b([A-Z]{2}\d{3})\s+(\d)\b", r"\1\2", value)
    value = re.sub(r"\b([A-Z][a-z]{2,3})\s+(\d)\s+(\d)\s+(202[67])\b", r"\1 \2\3 \4", value)
    return value.strip()


def clean_line(value: str) -> str:
    return clean_text(value)


def normalize_token(value: str) -> str | None:
    value = value.strip()
    if value in DASHES:
        return None
    if value == "Unlimited":
        return "Unlimited"
    return value


def parse_int_token(value: str | None) -> int | None:
    if value is None or value == "Unlimited":
        return None
    return int(value)


def infer_species(page_heading: str) -> str:
    heading = page_heading.upper()
    if "PRONGHORN" in heading:
        return "Pronghorn"
    if "ELK" in heading:
        return "Elk"
    if "DEER" in heading:
        return "Deer"
    if "MOOSE" in heading:
        return "Moose"
    if "MOUNTAIN GOAT" in heading:
        return "Mountain Goat"
    if "BISON" in heading:
        return "Bison"
    if "DESERT BIGHORN SHEEP" in heading:
        return "Desert Bighorn Sheep"
    if "ROCKY MOUNTAIN BIGHORN SHEEP" in heading or "EWE ROCKY MOUNTAIN BIGHORN SHEEP" in heading:
        return "Rocky Mountain Bighorn Sheep"
    return page_heading.title()


def infer_species_from_code(hunt_code: str) -> str:
    prefix = hunt_code[:2].upper()
    mapping = {
        "BI": "Bison",
        "DB": "Deer",
        "DA": "Deer",
        "DS": "Desert Bighorn Sheep",
        "EA": "Elk",
        "EB": "Elk",
        "EL": "Elk",
        "GO": "Mountain Goat",
        "LO": "Elk",
        "LP": "Pronghorn",
        "LD": "Deer",
        "MA": "Moose",
        "MB": "Moose",
        "PB": "Pronghorn",
        "PD": "Pronghorn",
        "RE": "Rocky Mountain Bighorn Sheep",
        "RS": "Rocky Mountain Bighorn Sheep",
    }
    return mapping.get(prefix, "")


def infer_access_class(page_heading: str, section_title: str, hunt_name: str) -> str:
    text = f"{page_heading} {section_title} {hunt_name}".upper()
    if "PRIVATE-LANDS-ONLY" in text or "PRIVATE LANDS ONLY" in text:
        return "Private"
    if "CWMU" in text:
        return "CWMU"
    return "Public"


def choose_weapon(prefix: str) -> tuple[str | None, str]:
    for option in WEAPON_OPTIONS:
        if prefix.endswith(option):
            remaining = prefix[: -len(option)].strip(" ,")
            return option.replace("Hunter's", "Hunter's"), remaining
    return None, prefix.strip()


def choose_sex_type(value: str) -> tuple[str | None, str]:
    for option in SEX_TYPE_OPTIONS:
        if value.startswith(option):
            remaining = value[len(option) :].strip()
            normalized = option.replace("Hunter ’s Choice", "Hunter's Choice")
            return normalized, remaining
    return None, value


def parse_split_numbers(values: list[str]) -> dict[str, object]:
    return {
        "resident_permits_prior_raw": normalize_token(values[0]),
        "nonresident_permits_prior_raw": normalize_token(values[1]),
        "total_permits_prior_raw": normalize_token(values[2]),
        "resident_permits_raw": normalize_token(values[3]),
        "nonresident_permits_raw": normalize_token(values[4]),
        "total_permits_raw": normalize_token(values[5]),
        "permit_split_type": "resident_nonresident",
    }


def parse_total_only_numbers(values: list[str]) -> dict[str, object]:
    current_raw = normalize_token(values[1])
    split_type = "unlimited_total_only" if current_raw == "Unlimited" else "total_only"
    return {
        "resident_permits_prior_raw": None,
        "nonresident_permits_prior_raw": None,
        "total_permits_prior_raw": normalize_token(values[0]),
        "resident_permits_raw": None,
        "nonresident_permits_raw": None,
        "total_permits_raw": current_raw,
        "permit_split_type": split_type,
    }


def base_row(context: SectionContext, hunt_code: str, hunt_name: str, weapon: str | None, sex_type: str | None) -> dict[str, object]:
    access_class = infer_access_class(context.page_heading, context.section_title, hunt_name)
    return {
        "recommendation_year": 2026,
        "hunt_code": hunt_code,
        "species": infer_species(context.page_heading),
        "permit_category": context.page_heading,
        "section_title": context.section_title,
        "hunt_name": hunt_name,
        "weapon": weapon or "",
        "sex_type": sex_type or "",
        "access_class": access_class,
        "source_file": PDF_PATH.name,
        "source_page_number": context.page_number,
        "source_type": "rac_recommended_permits",
        "source_authority_level": "recommended",
    }


def parse_context_weapon_split(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<weapon>.+?)\s+(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<n1>\S+)\s+(?P<n2>\S+)\s+(?P<n3>\S+)\s+(?P<n4>\S+)\s+(?P<n5>\S+)\s+(?P<n6>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed context/weapon split row: {line}")
    row = base_row(context, match.group("hunt_code"), context.section_title, clean_text(match.group("weapon")), None)
    row.update(parse_split_numbers([match.group(f"n{i}") for i in range(1, 7)]))
    return row


def parse_hunt_weapon_split(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<prefix>.+?)\s+(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<n1>\S+)\s+(?P<n2>\S+)\s+(?P<n3>\S+)\s+(?P<n4>\S+)\s+(?P<n5>\S+)\s+(?P<n6>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed hunt/weapon split row: {line}")
    weapon, hunt_name = choose_weapon(clean_text(match.group("prefix")))
    row = base_row(context, match.group("hunt_code"), hunt_name or context.section_title, weapon, None)
    row.update(parse_split_numbers([match.group(f"n{i}") for i in range(1, 7)]))
    return row


def parse_hunt_code_split_no_weapon(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<hunt_name>.+?)\s+(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<n1>\S+)\s+(?P<n2>\S+)\s+(?P<n3>\S+)\s+(?P<n4>\S+)\s+(?P<n5>\S+)\s+(?P<n6>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed hunt/code split row: {line}")
    row = base_row(context, match.group("hunt_code"), clean_text(match.group("hunt_name")), None, None)
    row.update(parse_split_numbers([match.group(f"n{i}") for i in range(1, 7)]))
    return row


def parse_split_with_sex(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<prefix>.+?)\s+(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<sex_text>.+?)\s+(?P<n1>\S+)\s+(?P<n2>\S+)\s+(?P<n3>\S+)\s+(?P<n4>\S+)\s+(?P<n5>\S+)\s+(?P<n6>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed split/sex row: {line}")
    weapon, hunt_name = choose_weapon(clean_text(match.group("prefix")))
    sex_type, trailing = choose_sex_type(clean_text(match.group("sex_text")))
    if trailing:
        raise ValueError(f"Unexpected trailing text after sex type: {line}")
    row = base_row(context, match.group("hunt_code"), hunt_name, weapon, sex_type)
    row.update(parse_split_numbers([match.group(f"n{i}") for i in range(1, 7)]))
    return row


def parse_code_first_split(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<body>.+?)\s+(?P<n1>\S+)\s+(?P<n2>\S+)\s+(?P<n3>\S+)\s+(?P<n4>\S+)\s+(?P<n5>\S+)\s+(?P<n6>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed code-first split row: {line}")
    body = clean_text(match.group("body"))
    season_match = re.search(r"\b[A-Z][a-z]{2,4}\s+\d{2}\s+202[67]\s+-\s+[A-Z][a-z]{2,4}\s+\d{2}\s+202[67]\b", body)
    if not season_match:
        raise ValueError(f"Unable to isolate season dates in row: {line}")
    left = body[: season_match.start()].strip()
    weapon, hunt_name = choose_weapon(left)
    row = base_row(context, match.group("hunt_code"), hunt_name, weapon, None)
    row.update(parse_split_numbers([match.group(f"n{i}") for i in range(1, 7)]))
    return row


def parse_code_first_total_only(context: SectionContext, line: str) -> dict[str, object]:
    match = re.match(
        r"^(?P<hunt_code>[A-Z]{2}\d{4})\s+(?P<body>.+?)\s+(?P<n1>\S+)\s+(?P<n2>\S+)$",
        line,
    )
    if not match:
        raise ValueError(f"Malformed code-first total-only row: {line}")
    body = clean_text(match.group("body"))
    season_match = re.search(r"\b[A-Z][a-z]{2,4}\s+\d{2}\s+202[67]\s+-\s+[A-Z][a-z]{2,4}\s+\d{2}\s+202[67]\b", body)
    if not season_match:
        raise ValueError(f"Unable to isolate season dates in row: {line}")
    left = body[: season_match.start()].strip()
    weapon, hunt_name = choose_weapon(left)
    row = base_row(context, match.group("hunt_code"), hunt_name, weapon, None)
    row.update(parse_total_only_numbers([match.group("n1"), match.group("n2")]))
    return row


def detect_header_type(header_line: str) -> str:
    header_line = clean_text(header_line)
    if "Sex Type" in header_line and header_line.startswith("Hunt Name Weapon Hunt Number"):
        return "split_with_sex"
    if header_line.startswith("Weapon Hunt Number"):
        return "context_weapon_split"
    if header_line.startswith("Hunt Name Weapon Hunt Number"):
        return "hunt_weapon_split"
    if header_line.startswith("Hunt Name Hunt Number"):
        return "hunt_code_split_no_weapon"
    if header_line.startswith("Hunt Number Hunt Name Weapon") and "Res Non" in header_line:
        return "code_first_split"
    if header_line.startswith("Hunt Number Hunt Name Weapon") and "2025 2026" in header_line:
        return "code_first_total_only"
    raise ValueError(f"Unsupported header line: {header_line}")


def looks_like_page_heading(line: str) -> bool:
    return clean_text(line).startswith("2026 ")


def has_year_marker(line: str) -> bool:
    return bool(YEAR_MARKER_RE.search(clean_text(line)))


def build_section_title(lines: list[str], index: int) -> tuple[str, int]:
    line = clean_text(lines[index])
    marker_match = YEAR_MARKER_RE.search(line)
    if marker_match:
        prefix = clean_text(line[: marker_match.start()])
        if prefix:
            prev = clean_text(lines[index - 1]) if index > 0 else ""
            if prev and not looks_like_page_heading(prev) and "Grand Total" not in prev and "Hunt Name" not in prev and "Weapon" not in prev:
                if len(prefix) <= 4:
                    return clean_text(f"{prev} {prefix}"), index
            return prefix, index
        for back_index in range(index - 1, max(-1, index - 3), -1):
            candidate = clean_text(lines[back_index])
            if not candidate:
                continue
            if looks_like_page_heading(candidate):
                return candidate, back_index
            if "Grand Total" in candidate or "Hunt Name" in candidate or "Weapon" in candidate or "Hunt Number" in candidate:
                continue
            return candidate, back_index
    raise ValueError(f"Unable to build section title from line: {line}")


def is_header_line(line: str) -> bool:
    line = clean_text(line)
    return "Hunt Number" in line and ("Weapon" in line or "Hunt Name" in line)


def build_header_candidate(lines: list[str], index: int) -> tuple[str, int]:
    pieces = [clean_text(lines[index])]
    consumed = 1
    for next_index in range(index + 1, min(len(lines), index + 4)):
        candidate = clean_text(lines[next_index])
        if not candidate:
            continue
        if looks_like_page_heading(candidate) or has_year_marker(candidate) or candidate.startswith("Grand Total"):
            break
        combined_so_far = clean_text(" ".join(pieces))
        if "Hunt Number" in combined_so_far and HUNT_CODE_RE.search(candidate):
            break
        if any(candidate.startswith(option) for option in WEAPON_OPTIONS) or re.match(r"^[A-Z]{2}\d{4}\b", candidate):
            break
        pieces.append(candidate)
        consumed += 1
    return clean_text(" ".join(pieces)), consumed


def is_row_start(line: str, header_type: str) -> bool:
    line = clean_text(line)
    if not line or line.startswith("Grand Total") or has_year_marker(line) or line.startswith("*"):
        return False
    if header_type == "context_weapon_split":
        return any(line.startswith(option) for option in WEAPON_OPTIONS)
    if header_type == "code_first_split" or header_type == "code_first_total_only":
        return bool(re.match(r"^[A-Z]{2}\d{4}\b", line))
    return bool(HUNT_CODE_RE.search(line))


def ends_with_enough_tokens(line: str, header_type: str) -> bool:
    line = clean_text(line)
    if header_type in {"context_weapon_split", "hunt_weapon_split", "hunt_code_split_no_weapon", "split_with_sex", "code_first_split"}:
        return bool(re.search(r"(?:\S+\s+){5}\S+$", line))
    if header_type == "code_first_total_only":
        return bool(re.search(r"\S+\s+\S+$", line))
    return False


def parse_row(context: SectionContext, line: str) -> dict[str, object]:
    line = clean_text(line)
    if context.header_type == "context_weapon_split":
        return parse_context_weapon_split(context, line)
    if context.header_type == "hunt_weapon_split":
        return parse_hunt_weapon_split(context, line)
    if context.header_type == "hunt_code_split_no_weapon":
        return parse_hunt_code_split_no_weapon(context, line)
    if context.header_type == "split_with_sex":
        return parse_split_with_sex(context, line)
    if context.header_type == "code_first_split":
        return parse_code_first_split(context, line)
    if context.header_type == "code_first_total_only":
        return parse_code_first_total_only(context, line)
    raise ValueError(f"Unsupported header type: {context.header_type}")


def extract_rows(reader: PdfReader) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    section_records: list[dict[str, object]] = []
    rows: list[dict[str, object]] = []

    for page_number, page in enumerate(reader.pages, start=1):
        raw_text = page.extract_text() or ""
        lines = [clean_line(line) for line in raw_text.splitlines() if clean_line(line)]
        if not any(has_year_marker(line) for line in lines):
            continue

        page_heading = next((line for line in lines if looks_like_page_heading(line)), "")
        context: SectionContext | None = None
        current_row_parts: list[str] = []
        skip_indexes: set[int] = set()

        def finalize_current_row() -> None:
            nonlocal current_row_parts, context
            if not current_row_parts or context is None:
                current_row_parts = []
                return
            row_text = clean_text(" ".join(current_row_parts))
            parsed = parse_row(context, row_text)
            rows.append(parsed)
            section_records.append(
                {
                    "page_number": page_number,
                    "page_heading": page_heading,
                    "section_title": context.section_title,
                    "header_type": context.header_type,
                    "row_text": row_text,
                    "hunt_code": parsed["hunt_code"],
                }
            )
            current_row_parts = []

        for index, line in enumerate(lines):
            if index in skip_indexes:
                continue
            if has_year_marker(line):
                finalize_current_row()
                section_title, _ = build_section_title(lines, index)
                context = SectionContext(
                    page_number=page_number,
                    page_heading=page_heading,
                    section_title=section_title,
                    header_type="",
                )
                continue

            if context is None:
                continue

            header_candidate, consumed = build_header_candidate(lines, index)
            if is_header_line(header_candidate):
                context.header_type = detect_header_type(header_candidate)
                for skip_index in range(index + 1, index + consumed):
                    skip_indexes.add(skip_index)
                continue

            if not context.header_type:
                continue

            if line.startswith("Grand Total") or line.startswith("*This unit"):
                finalize_current_row()
                continue

            if is_row_start(line, context.header_type):
                finalize_current_row()
                current_row_parts = [line]
                if ends_with_enough_tokens(line, context.header_type):
                    finalize_current_row()
                continue

            if current_row_parts:
                current_row_parts.append(line)
                if ends_with_enough_tokens(clean_text(" ".join(current_row_parts)), context.header_type):
                    finalize_current_row()

        finalize_current_row()

    return section_records, rows


def enrich_numeric_fields(rows: list[dict[str, object]]) -> None:
    for row in rows:
        row["resident_permits_prior"] = parse_int_token(row["resident_permits_prior_raw"])
        row["nonresident_permits_prior"] = parse_int_token(row["nonresident_permits_prior_raw"])
        row["total_permits_prior"] = parse_int_token(row["total_permits_prior_raw"])
        row["resident_permits"] = parse_int_token(row["resident_permits_raw"])
        row["nonresident_permits"] = parse_int_token(row["nonresident_permits_raw"])
        row["total_permits"] = parse_int_token(row["total_permits_raw"])


def dedupe_rows(rows: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[str]]:
    deduped: list[dict[str, object]] = []
    seen_by_code: dict[str, dict[str, object]] = {}
    conflicts: list[str] = []

    compare_fields = [
        "species",
        "hunt_name",
        "weapon",
        "sex_type",
        "access_class",
        "permit_split_type",
        "resident_permits_prior_raw",
        "nonresident_permits_prior_raw",
        "total_permits_prior_raw",
        "resident_permits_raw",
        "nonresident_permits_raw",
        "total_permits_raw",
    ]

    for row in rows:
        hunt_code = str(row["hunt_code"])
        prior = seen_by_code.get(hunt_code)
        if prior is None:
            seen_by_code[hunt_code] = row
            deduped.append(row)
            continue

        same = all(str(prior.get(field, "")) == str(row.get(field, "")) for field in compare_fields)
        if same:
            continue

        conflicts.append(hunt_code)

    return deduped, sorted(set(conflicts))


def write_jsonl(records: list[dict[str, object]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=True))
            handle.write("\n")


def write_csv(rows: list[dict[str, object]], path: Path) -> None:
    fieldnames = [
        "recommendation_year",
        "hunt_code",
        "species",
        "permit_category",
        "section_title",
        "hunt_name",
        "weapon",
        "sex_type",
        "access_class",
        "permit_split_type",
        "resident_permits_prior_raw",
        "nonresident_permits_prior_raw",
        "total_permits_prior_raw",
        "resident_permits_raw",
        "nonresident_permits_raw",
        "total_permits_raw",
        "resident_permits_prior",
        "nonresident_permits_prior",
        "total_permits_prior",
        "resident_permits",
        "nonresident_permits",
        "total_permits",
        "source_type",
        "source_authority_level",
        "source_file",
        "source_page_number",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def validate(rows: list[dict[str, object]]) -> dict[str, object]:
    hunt_codes = [str(row["hunt_code"]) for row in rows]
    numeric_checks = {
        "resident_permits": sum(1 for row in rows if row["resident_permits"] is not None),
        "nonresident_permits": sum(1 for row in rows if row["nonresident_permits"] is not None),
        "total_permits": sum(1 for row in rows if row["total_permits"] is not None),
    }
    split_counts: dict[str, int] = {}
    for row in rows:
        split_counts[str(row["permit_split_type"])] = split_counts.get(str(row["permit_split_type"]), 0) + 1

    malformed = []
    for row in rows:
        if row["permit_split_type"] == "resident_nonresident":
            if row["resident_permits"] is None or row["nonresident_permits"] is None or row["total_permits"] is None:
                malformed.append(str(row["hunt_code"]))

    return {
        "row_count": len(rows),
        "distinct_hunt_code_count": len(set(hunt_codes)),
        "duplicate_hunt_codes": [],
        "numeric_field_non_null_counts": numeric_checks,
        "permit_split_type_counts": split_counts,
        "malformed_hunt_codes": sorted(set(malformed)),
    }


def parse_labeled_permit(value: object, label: str) -> int | None:
    if value is None:
        return None
    text = clean_text(str(value))
    match = re.fullmatch(rf"{label}\s*:\s*(\d+)", text, re.IGNORECASE)
    if not match:
        empty_match = re.fullmatch(rf"{label}\s*:\s*", text, re.IGNORECASE)
        return 0 if empty_match else None
    return int(match.group(1))


def load_buck_deer_workbook_rows() -> dict[str, dict[str, object]]:
    workbook = load_workbook(DEER_WORKBOOK_PATH, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    deer_rows: dict[str, dict[str, object]] = {}
    buffered_rows = list(sheet.iter_rows(values_only=True))

    for index, row in enumerate(buffered_rows):
        hunt_code = row[1]
        if not isinstance(hunt_code, str):
            continue

        hunt_code = hunt_code.strip()
        if not re.fullmatch(r"DB\d{4}", hunt_code):
            continue

        resident_permits = parse_labeled_permit(row[7], "Res")
        total_only = parse_labeled_permit(row[7], "Total")
        next_row = buffered_rows[index + 1] if index + 1 < len(buffered_rows) else (None,) * 8
        nonresident_permits = parse_labeled_permit(next_row[7], "NonRes")

        if total_only is not None:
            deer_rows[hunt_code] = {
                "recommendation_year": 2026,
                "hunt_code": hunt_code,
                "species": "Deer",
                "permit_category": "2026 Buck Deer Permit Recommendations",
                "section_title": "Buck Deer",
                "hunt_name": clean_text(str(row[0] or "")),
                "weapon": clean_text(str(row[4] or "")),
                "sex_type": "",
                "access_class": "Public",
                "permit_split_type": "total_only",
                "resident_permits_prior_raw": None,
                "nonresident_permits_prior_raw": None,
                "total_permits_prior_raw": None,
                "resident_permits_raw": None,
                "nonresident_permits_raw": None,
                "total_permits_raw": str(total_only),
                "resident_permits_prior": None,
                "nonresident_permits_prior": None,
                "total_permits_prior": None,
                "resident_permits": None,
                "nonresident_permits": None,
                "total_permits": total_only,
                "source_type": "buck_deer_permits_workbook",
                "source_authority_level": "updated_current_permits",
                "source_file": DEER_WORKBOOK_PATH.name,
                "source_page_number": None,
            }
            continue

        if resident_permits is None and nonresident_permits is None:
            continue

        total_permits = None
        if resident_permits is not None or nonresident_permits is not None:
            total_permits = (resident_permits or 0) + (nonresident_permits or 0)

        deer_rows[hunt_code] = {
            "recommendation_year": 2026,
            "hunt_code": hunt_code,
            "species": "Deer",
            "permit_category": "2026 Buck Deer Permit Recommendations",
            "section_title": "Buck Deer",
            "hunt_name": clean_text(str(row[0] or "")),
            "weapon": clean_text(str(row[4] or "")),
            "sex_type": "",
            "access_class": "Public",
            "permit_split_type": "resident_nonresident",
            "resident_permits_prior_raw": None,
            "nonresident_permits_prior_raw": None,
            "total_permits_prior_raw": None,
            "resident_permits_raw": str(resident_permits) if resident_permits is not None else None,
            "nonresident_permits_raw": str(nonresident_permits) if nonresident_permits is not None else None,
            "total_permits_raw": str(total_permits) if total_permits is not None else None,
            "resident_permits_prior": None,
            "nonresident_permits_prior": None,
            "total_permits_prior": None,
            "resident_permits": resident_permits,
            "nonresident_permits": nonresident_permits,
            "total_permits": total_permits,
            "source_type": "buck_deer_permits_workbook",
            "source_authority_level": "updated_current_permits",
            "source_file": DEER_WORKBOOK_PATH.name,
            "source_page_number": None,
        }

    return deer_rows


def infer_access_class_from_workbook(hunt_type: str, hunt_name: str) -> str:
    text = clean_text(f"{hunt_type} {hunt_name}").upper()
    if "PRIVATE LAND ONLY" in text or "PRIVATE-LAND-ONLY" in text:
        return "Private"
    if "CWMU" in text:
        return "CWMU"
    return "Public"


def parse_workbook_permit_text(value: object) -> tuple[int | None, int | None, int | None]:
    text = clean_text(str(value or ""))
    return (
        parse_labeled_permit(text, "Res"),
        parse_labeled_permit(text, "NonRes"),
        parse_labeled_permit(text, "Total"),
    )


def parse_plain_int(value: str) -> int | None:
    text = clean_text(value)
    return int(text) if re.fullmatch(r"\d+", text) else None


def build_workbook_row(
    *,
    path: Path,
    hunt_code: str,
    species: str,
    hunt_name: str,
    weapon: str,
    sex_type: str,
    hunt_type: str,
    total_permits: int | None,
    resident_permits: int | None,
    nonresident_permits: int | None,
    permit_split_type: str,
    permit_category: str,
    section_title: str,
) -> dict[str, object]:
    return {
        "recommendation_year": 2026,
        "hunt_code": hunt_code,
        "species": species,
        "permit_category": permit_category,
        "section_title": section_title,
        "hunt_name": hunt_name,
        "weapon": weapon,
        "sex_type": sex_type,
        "access_class": infer_access_class_from_workbook(hunt_type, hunt_name),
        "permit_split_type": permit_split_type,
        "resident_permits_prior_raw": None,
        "nonresident_permits_prior_raw": None,
        "total_permits_prior_raw": None,
        "resident_permits_raw": str(resident_permits) if resident_permits is not None else None,
        "nonresident_permits_raw": str(nonresident_permits) if nonresident_permits is not None else None,
        "total_permits_raw": str(total_permits) if total_permits is not None else None,
        "resident_permits_prior": None,
        "nonresident_permits_prior": None,
        "total_permits_prior": None,
        "resident_permits": resident_permits,
        "nonresident_permits": nonresident_permits,
        "total_permits": total_permits,
        "source_type": path.stem.lower().replace(" ", "_").replace(".", "").replace("'", "") + "_workbook",
        "source_authority_level": "updated_current_permits",
        "source_file": path.name,
        "source_page_number": None,
    }


def load_rac_workbook_overlay_rows() -> dict[str, dict[str, object]]:
    parsed_rows: dict[str, dict[str, object]] = {}
    workbook_paths = [RAC_WORKBOOK_PATH, RAC_WORKBOOK_ALT_PATH]

    for path in workbook_paths:
        if not path.exists():
            continue
        workbook = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                values = [clean_text(str(value)) for value in row if value is not None and clean_text(str(value))]
                if not values:
                    continue

                code_index = next((index for index, value in enumerate(values) if re.fullmatch(r"[A-Z]{2}\d{4}", value)), None)
                if code_index is None:
                    continue

                hunt_code = values[code_index]
                tail = values[code_index + 1 :]

                if len(tail) == 5:
                    hunt_name = values[code_index + 1]
                    weapon = values[code_index + 2]
                    prior_total = parse_plain_int(values[code_index + 4])
                    current_total = parse_plain_int(values[code_index + 5]) if len(values) > code_index + 5 else None
                    if current_total is None:
                        continue
                    parsed_rows[hunt_code] = build_workbook_row(
                        path=path,
                        hunt_code=hunt_code,
                        species=infer_species_from_code(hunt_code),
                        hunt_name=hunt_name,
                        weapon=weapon,
                        sex_type="",
                        hunt_type="",
                        total_permits=current_total,
                        resident_permits=None,
                        nonresident_permits=None,
                        permit_split_type="total_only",
                        permit_category="2026 RAC Workbook Permit Recommendations",
                        section_title=sheet_name,
                    )
                    if prior_total is not None:
                        parsed_rows[hunt_code]["total_permits_prior_raw"] = str(prior_total)
                        parsed_rows[hunt_code]["total_permits_prior"] = prior_total
                    continue

                if len(values) >= code_index + 7:
                    trailing = values[-6:]
                    if all(parse_plain_int(value) is not None for value in trailing):
                        resident_prior = parse_plain_int(trailing[0])
                        nonresident_prior = parse_plain_int(trailing[1])
                        total_prior = parse_plain_int(trailing[2])
                        resident_current = parse_plain_int(trailing[3])
                        nonresident_current = parse_plain_int(trailing[4])
                        total_current = parse_plain_int(trailing[5])
                        if total_current is None:
                            continue

                        before = values[: code_index] + values[code_index + 1 : -6]
                        hunt_name = before[0] if before else ""
                        weapon = next((value for value in before if value in WEAPON_OPTIONS), "")
                        sex_type = next((value for value in before if value in SEX_TYPE_OPTIONS), "")

                        parsed_rows[hunt_code] = build_workbook_row(
                            path=path,
                            hunt_code=hunt_code,
                            species=infer_species_from_code(hunt_code),
                            hunt_name=hunt_name,
                            weapon=weapon,
                            sex_type=sex_type,
                            hunt_type="",
                            total_permits=total_current,
                            resident_permits=resident_current,
                            nonresident_permits=nonresident_current,
                            permit_split_type="resident_nonresident",
                            permit_category="2026 RAC Workbook Permit Recommendations",
                            section_title=sheet_name,
                        )
                        parsed_rows[hunt_code]["resident_permits_prior_raw"] = str(resident_prior) if resident_prior is not None else None
                        parsed_rows[hunt_code]["nonresident_permits_prior_raw"] = str(nonresident_prior) if nonresident_prior is not None else None
                        parsed_rows[hunt_code]["total_permits_prior_raw"] = str(total_prior) if total_prior is not None else None
                        parsed_rows[hunt_code]["resident_permits_prior"] = resident_prior
                        parsed_rows[hunt_code]["nonresident_permits_prior"] = nonresident_prior
                        parsed_rows[hunt_code]["total_permits_prior"] = total_prior

    return parsed_rows


def load_split_workbook_rows(path: Path, code_pattern: str, species: str, permit_category: str, section_title: str) -> dict[str, dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    buffered_rows = list(sheet.iter_rows(values_only=True))
    parsed_rows: dict[str, dict[str, object]] = {}

    for index, row in enumerate(buffered_rows):
        if len(row) < 8:
            continue
        hunt_code = row[1]
        if not isinstance(hunt_code, str) or not re.fullmatch(code_pattern, hunt_code.strip()):
            continue

        resident, nonresident, total_only = parse_workbook_permit_text(row[7])
        if total_only is not None:
            parsed_rows[hunt_code.strip()] = build_workbook_row(
                path=path,
                hunt_code=hunt_code.strip(),
                species=species,
                hunt_name=clean_text(str(row[0] or "")),
                weapon=clean_text(str(row[4] or "")),
                sex_type=clean_text(str(row[2] or "")),
                hunt_type=clean_text(str(row[5] or "")),
                total_permits=total_only,
                resident_permits=None,
                nonresident_permits=None,
                permit_split_type="total_only",
                permit_category=permit_category,
                section_title=section_title,
            )
            continue

        next_row = buffered_rows[index + 1] if index + 1 < len(buffered_rows) else (None,) * 8
        next_nonresident = parse_labeled_permit(next_row[7] if len(next_row) > 7 else None, "NonRes")
        if nonresident is None:
            nonresident = next_nonresident
        if resident is None and nonresident is None:
            continue

        total = total_only if total_only is not None else (resident or 0) + (nonresident or 0)
        parsed_rows[hunt_code.strip()] = build_workbook_row(
            path=path,
            hunt_code=hunt_code.strip(),
            species=species,
            hunt_name=clean_text(str(row[0] or "")),
            weapon=clean_text(str(row[4] or "")),
            sex_type=clean_text(str(row[2] or "")),
            hunt_type=clean_text(str(row[5] or "")),
            total_permits=total,
            resident_permits=resident,
            nonresident_permits=nonresident,
            permit_split_type="resident_nonresident",
            permit_category=permit_category,
            section_title=section_title,
        )

    return parsed_rows


def load_total_only_workbook_rows(path: Path, code_pattern: str, species: str, permit_category: str, section_title: str) -> dict[str, dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    parsed_rows: dict[str, dict[str, object]] = {}

    for row in sheet.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        hunt_code = row[1]
        if not isinstance(hunt_code, str) or not re.fullmatch(code_pattern, hunt_code.strip()):
            continue
        _, _, total = parse_workbook_permit_text(row[7])
        if total is None:
            continue
        parsed_rows[hunt_code.strip()] = build_workbook_row(
            path=path,
            hunt_code=hunt_code.strip(),
            species=species,
            hunt_name=clean_text(str(row[0] or "")),
            weapon=clean_text(str(row[4] or "")),
            sex_type=clean_text(str(row[2] or "")),
            hunt_type=clean_text(str(row[5] or "")),
            total_permits=total,
            resident_permits=None,
            nonresident_permits=None,
            permit_split_type="total_only",
            permit_category=permit_category,
            section_title=section_title,
        )

    return parsed_rows


def load_mixed_antlerless_deer_rows() -> dict[str, dict[str, object]]:
    workbook = load_workbook(ANTLERLESS_DEER_WORKBOOK_PATH, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    buffered_rows = list(sheet.iter_rows(values_only=True))
    parsed_rows: dict[str, dict[str, object]] = {}

    for index, row in enumerate(buffered_rows):
        if len(row) < 8:
            continue
        hunt_code = row[1]
        if not isinstance(hunt_code, str) or not re.fullmatch(r"DA\d{4}", hunt_code.strip()):
            continue
        resident, nonresident, total = parse_workbook_permit_text(row[7])
        if total is not None:
            parsed_rows[hunt_code.strip()] = build_workbook_row(
                path=ANTLERLESS_DEER_WORKBOOK_PATH,
                hunt_code=hunt_code.strip(),
                species="Deer",
                hunt_name=clean_text(str(row[0] or "")),
                weapon=clean_text(str(row[4] or "")),
                sex_type=clean_text(str(row[2] or "")),
                hunt_type=clean_text(str(row[5] or "")),
                total_permits=total,
                resident_permits=None,
                nonresident_permits=None,
                permit_split_type="total_only",
                permit_category="2026 Antlerless Deer Permit Recommendations",
                section_title="Antlerless Deer",
            )
            continue

        next_row = buffered_rows[index + 1] if index + 1 < len(buffered_rows) else (None,) * 8
        next_nonresident = parse_labeled_permit(next_row[7] if len(next_row) > 7 else None, "NonRes")
        if nonresident is None:
            nonresident = next_nonresident
        if resident is None and nonresident is None:
            continue
        parsed_rows[hunt_code.strip()] = build_workbook_row(
            path=ANTLERLESS_DEER_WORKBOOK_PATH,
            hunt_code=hunt_code.strip(),
            species="Deer",
            hunt_name=clean_text(str(row[0] or "")),
            weapon=clean_text(str(row[4] or "")),
            sex_type=clean_text(str(row[2] or "")),
            hunt_type=clean_text(str(row[5] or "")),
            total_permits=(resident or 0) + (nonresident or 0),
            resident_permits=resident,
            nonresident_permits=nonresident,
            permit_split_type="resident_nonresident",
            permit_category="2026 Antlerless Deer Permit Recommendations",
            section_title="Antlerless Deer",
        )

    return parsed_rows


def load_turkey_workbook_rows() -> dict[str, dict[str, object]]:
    workbook = load_workbook(TURKEY_WORKBOOK_PATH, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    turkey_rows: dict[str, dict[str, object]] = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        hunt_code = row[1]
        if not isinstance(hunt_code, str):
            continue

        hunt_code = hunt_code.strip()
        if not re.fullmatch(r"TK\d{4}", hunt_code):
            continue

        total_match = re.fullmatch(r"Total\s*:\s*(\d+)", clean_text(str(row[7] or "")), re.IGNORECASE)
        if not total_match:
            continue

        total_permits = int(total_match.group(1))
        access_class = "CWMU" if clean_text(str(row[5] or "")).upper() == "CWMU" else "Public"
        turkey_rows[hunt_code] = {
            "recommendation_year": 2026,
            "hunt_code": hunt_code,
            "species": "Turkey",
            "permit_category": "2026 Bearded Turkey Permit Recommendations",
            "section_title": "Turkey Bearded",
            "hunt_name": clean_text(str(row[0] or "")),
            "weapon": clean_text(str(row[4] or "")),
            "sex_type": clean_text(str(row[2] or "")),
            "access_class": access_class,
            "permit_split_type": "total_only",
            "resident_permits_prior_raw": None,
            "nonresident_permits_prior_raw": None,
            "total_permits_prior_raw": None,
            "resident_permits_raw": None,
            "nonresident_permits_raw": None,
            "total_permits_raw": str(total_permits),
            "resident_permits_prior": None,
            "nonresident_permits_prior": None,
            "total_permits_prior": None,
            "resident_permits": None,
            "nonresident_permits": None,
            "total_permits": total_permits,
            "source_type": "turkey_bearded_workbook",
            "source_authority_level": "updated_current_permits",
            "source_file": TURKEY_WORKBOOK_PATH.name,
            "source_page_number": None,
        }

    return turkey_rows


def apply_buck_deer_workbook_overrides(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    workbook_rows = load_buck_deer_workbook_rows()
    if not workbook_rows:
        return rows

    by_code = {str(row["hunt_code"]): dict(row) for row in rows}
    for hunt_code, workbook_row in workbook_rows.items():
        prior_row = by_code.get(hunt_code)
        if prior_row:
            workbook_row["resident_permits_prior_raw"] = prior_row.get("resident_permits_prior_raw")
            workbook_row["nonresident_permits_prior_raw"] = prior_row.get("nonresident_permits_prior_raw")
            workbook_row["total_permits_prior_raw"] = prior_row.get("total_permits_prior_raw")
            workbook_row["resident_permits_prior"] = prior_row.get("resident_permits_prior")
            workbook_row["nonresident_permits_prior"] = prior_row.get("nonresident_permits_prior")
            workbook_row["total_permits_prior"] = prior_row.get("total_permits_prior")
            if prior_row.get("section_title"):
                workbook_row["section_title"] = prior_row["section_title"]
            if prior_row.get("permit_category"):
                workbook_row["permit_category"] = prior_row["permit_category"]
            if prior_row.get("access_class"):
                workbook_row["access_class"] = prior_row["access_class"]
        by_code[hunt_code] = workbook_row

    merged_rows = list(by_code.values())
    merged_rows.sort(key=lambda row: (str(row["hunt_code"]), "" if row["source_page_number"] is None else str(row["source_page_number"])))
    return merged_rows


def apply_workbook_overrides(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    by_code = {str(row["hunt_code"]): dict(row) for row in rows}

    workbook_sources: list[dict[str, dict[str, object]]] = [
        load_rac_workbook_overlay_rows(),
        load_buck_deer_workbook_rows(),
        load_split_workbook_rows(ANTLERLESS_ELK_WORKBOOK_PATH, r"EA\d{4}", "Elk", "2026 Antlerless Elk Permit Recommendations", "Antlerless Elk"),
        load_mixed_antlerless_deer_rows(),
        load_split_workbook_rows(BISON_BULL_WORKBOOK_PATH, r"BI\d{4}", "Bison", "2026 Bison Permit Recommendations", "Bison Bull"),
        load_split_workbook_rows(BISON_COW_WORKBOOK_PATH, r"BI\d{4}", "Bison", "2026 Bison Permit Recommendations", "Bison Cow"),
        load_split_workbook_rows(COW_MOOSE_WORKBOOK_PATH, r"MA\d{4}", "Moose", "2026 Antlerless Moose Permit Recommendations", "Cow Moose"),
        load_split_workbook_rows(BULL_MOOSE_WORKBOOK_PATH, r"MB\d{4}", "Moose", "2026 Bull Moose Permit Recommendations", "Bull Moose"),
        load_split_workbook_rows(MOUNTAIN_GOAT_WORKBOOK_PATH, r"GO\d{4}", "Mountain Goat", "2026 Mountain Goat Permit Recommendations", "Mountain Goat"),
        load_split_workbook_rows(PRONGHORN_BUCK_WORKBOOK_PATH, r"PB\d{4}", "Pronghorn", "2026 Buck Pronghorn Permit Recommendations", "Pronghorn Buck"),
        load_split_workbook_rows(PRONGHORN_DOE_WORKBOOK_PATH, r"PD\d{4}", "Pronghorn", "2026 Doe Pronghorn Permit Recommendations", "Pronghorn Doe"),
        load_split_workbook_rows(DESERT_SHEEP_RAM_WORKBOOK_PATH, r"DS\d{4}", "Desert Bighorn Sheep", "2026 Desert Bighorn Sheep Permit Recommendations", "Desert Sheep Ram"),
        load_split_workbook_rows(ROCKY_RAM_WORKBOOK_PATH, r"RS\d{4}", "Rocky Mountain Bighorn Sheep", "2026 Rocky Mountain Bighorn Sheep Permit Recommendations", "Rocky Ram"),
        load_split_workbook_rows(ROCKY_EWE_WORKBOOK_PATH, r"RE\d{4}", "Rocky Mountain Bighorn Sheep", "2026 Rocky Mountain Bighorn Sheep Permit Recommendations", "Rocky Ewe"),
        load_total_only_workbook_rows(TURKEY_EITHER_SEX_WORKBOOK_PATH, r"TK\d{4}", "Turkey", "2026 Either Sex Turkey Permit Recommendations", "Turkey Either Sex"),
        load_total_only_workbook_rows(DEER_HUNTERS_CHOICE_WORKBOOK_PATH, r"DB\d{4}", "Deer", "2026 Deer Hunter's Choice Permit Recommendations", "Deer Hunter's Choice"),
    ]

    for workbook_rows in workbook_sources:
        for hunt_code, workbook_row in workbook_rows.items():
            prior_row = by_code.get(hunt_code)
            if prior_row:
                workbook_row["resident_permits_prior_raw"] = prior_row.get("resident_permits_prior_raw")
                workbook_row["nonresident_permits_prior_raw"] = prior_row.get("nonresident_permits_prior_raw")
                workbook_row["total_permits_prior_raw"] = prior_row.get("total_permits_prior_raw")
                workbook_row["resident_permits_prior"] = prior_row.get("resident_permits_prior")
                workbook_row["nonresident_permits_prior"] = prior_row.get("nonresident_permits_prior")
                workbook_row["total_permits_prior"] = prior_row.get("total_permits_prior")
                if workbook_row["permit_split_type"] == "resident_nonresident":
                    if workbook_row["resident_permits"] is None and prior_row.get("resident_permits") is not None:
                        workbook_row["resident_permits"] = prior_row.get("resident_permits")
                        workbook_row["resident_permits_raw"] = prior_row.get("resident_permits_raw")
                    if workbook_row["nonresident_permits"] is None and prior_row.get("nonresident_permits") is not None:
                        workbook_row["nonresident_permits"] = prior_row.get("nonresident_permits")
                        workbook_row["nonresident_permits_raw"] = prior_row.get("nonresident_permits_raw")
                    if workbook_row["resident_permits"] is not None or workbook_row["nonresident_permits"] is not None:
                        workbook_row["total_permits"] = (workbook_row["resident_permits"] or 0) + (workbook_row["nonresident_permits"] or 0)
                        workbook_row["total_permits_raw"] = str(workbook_row["total_permits"])
            by_code[hunt_code] = workbook_row

    merged_rows = list(by_code.values())
    merged_rows.sort(key=lambda row: (str(row["hunt_code"]), "" if row["source_page_number"] is None else str(row["source_page_number"])))
    return merged_rows


def apply_turkey_workbook_overrides(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    workbook_rows = load_turkey_workbook_rows()
    if not workbook_rows:
        return rows

    by_code = {str(row["hunt_code"]): dict(row) for row in rows}
    for hunt_code, workbook_row in workbook_rows.items():
        prior_row = by_code.get(hunt_code)
        if prior_row:
            workbook_row["resident_permits_prior_raw"] = prior_row.get("resident_permits_prior_raw")
            workbook_row["nonresident_permits_prior_raw"] = prior_row.get("nonresident_permits_prior_raw")
            workbook_row["total_permits_prior_raw"] = prior_row.get("total_permits_prior_raw")
            workbook_row["resident_permits_prior"] = prior_row.get("resident_permits_prior")
            workbook_row["nonresident_permits_prior"] = prior_row.get("nonresident_permits_prior")
            workbook_row["total_permits_prior"] = prior_row.get("total_permits_prior")
            if prior_row.get("permit_category"):
                workbook_row["permit_category"] = prior_row["permit_category"]
            if prior_row.get("section_title"):
                workbook_row["section_title"] = prior_row["section_title"]
        by_code[hunt_code] = workbook_row

    merged_rows = list(by_code.values())
    merged_rows.sort(key=lambda row: (str(row["hunt_code"]), "" if row["source_page_number"] is None else str(row["source_page_number"])))
    return merged_rows


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    reader = PdfReader(str(PDF_PATH))
    section_records, rows = extract_rows(reader)
    enrich_numeric_fields(rows)
    rows, duplicate_conflicts = dedupe_rows(rows)
    rows = apply_buck_deer_workbook_overrides(rows)
    rows = apply_workbook_overrides(rows)
    rows = apply_turkey_workbook_overrides(rows)
    rows.sort(key=lambda row: (str(row["hunt_code"]), "" if row["source_page_number"] is None else str(row["source_page_number"])))

    write_jsonl(section_records, RAW_SECTION_PATH)
    write_csv(rows, CSV_PATH)

    validation = validate(rows)
    validation["duplicate_hunt_code_conflicts"] = duplicate_conflicts
    accepted_for_use = not duplicate_conflicts and not validation["malformed_hunt_codes"]
    validation["accepted_for_use"] = accepted_for_use

    status = {
        "dataset": "recommended_permits_2026",
        "source_file": str(PDF_PATH),
        "accepted_for_use": accepted_for_use,
        "authoritative_role": "current_recommended_permit_numbers",
        "published_output_path": str(PROCESSED_OUTPUT_PATH),
        "notes": [
            "This layer preserves resident and nonresident permit numbers when explicitly present in the RAC packet.",
            "Buck deer hunts are overlaid from the updated BUCK DEER workbook when that resource provides newer resident/nonresident permit counts.",
            "Additional workbook overlays are applied for antlerless elk, antlerless deer, bison, moose, mountain goat, pronghorn, sheep, turkey, and deer hunter's choice when those workbooks provide explicit permit counts.",
            "Turkey bearded hunts are overlaid from the Turkey workbook when it provides explicit total-only permit counts.",
            "Rows that only publish total permits remain total_only with resident/nonresident left null.",
            "This layer is separate from harvest and draw-history layers.",
        ],
    }
    acceptance = {
        "accepted_for_use": accepted_for_use,
        "row_count": validation["row_count"],
        "distinct_hunt_code_count": validation["distinct_hunt_code_count"],
        "duplicate_hunt_codes": validation["duplicate_hunt_codes"],
        "duplicate_hunt_code_conflicts": duplicate_conflicts,
        "malformed_hunt_codes": validation["malformed_hunt_codes"],
        "published_output_path": str(PROCESSED_OUTPUT_PATH),
    }
    source_manifest = {
        "source_file": str(PDF_PATH),
        "source_type": "utah_rac_packet",
        "recommendation_year": 2026,
        "authoritative_role": "recommended_permits",
        "supplemental_source_file": str(DEER_WORKBOOK_PATH),
        "supplemental_source_type": "buck_deer_permits_workbook",
        "supplemental_source_file_2": str(TURKEY_WORKBOOK_PATH),
        "supplemental_source_type_2": "turkey_bearded_workbook",
        "published_year_output_path": str(PROCESSED_YEAR_OUTPUT_PATH),
    }

    VALIDATION_PATH.write_text(json.dumps(validation, indent=2), encoding="utf-8")
    STATUS_PATH.write_text(json.dumps(status, indent=2), encoding="utf-8")
    ACCEPTANCE_PATH.write_text(json.dumps(acceptance, indent=2), encoding="utf-8")
    SOURCE_MANIFEST_PATH.write_text(json.dumps(source_manifest, indent=2), encoding="utf-8")

    if accepted_for_use:
        PROCESSED_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(CSV_PATH, PROCESSED_OUTPUT_PATH)
        PROCESSED_YEAR_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(CSV_PATH, PROCESSED_YEAR_OUTPUT_PATH)


if __name__ == "__main__":
    main()
